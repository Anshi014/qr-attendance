import os
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, g

from database_logic import (
    init_db, seed_students_from_excel, load_student_list,
    roll_exists, has_already_submitted, mark_attendance, get_student_name,
    add_user, delete_user, get_all_users, get_teachers, get_incharges,
    add_class, delete_class, get_all_classes, assign_incharge_to_class,
    add_subject, delete_subject, assign_teacher_to_subject,
    get_subjects_for_user, get_subject_by_id, get_all_subjects,
    authenticate_user, generate_report_for_subject,
    get_attendance_records, get_attendance_summary,
    get_students_by_class, get_all_students, add_student, delete_student,
)
from qr_generator import generate_qr, issue_device_token, consume_device_token, cleanup_old_tokens

app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY", "fallback-only-for-dev")

init_db()


def seed_all_classes():
    if not os.path.exists("student_list.xlsx"):
        return
    from database_logic import get_db
    conn = get_db()
    rows = conn.run("SELECT id FROM classes")
    class_ids = [row[0] for row in rows]
    conn.close()
    for cid in class_ids:
        seed_students_from_excel(class_id=cid)

seed_all_classes()


# ── Per-request DB connection (DEPRECATED — all DB ops now use database_logic) ──────────────────────────────────────────────

# def get_db():
#     if "db" not in g:
#         g.db = sqlite3.connect(DB, check_same_thread=False)
#         g.db.row_factory = sqlite3.Row
#     return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        db.close()


def ensure_indexes():
    """DEPRECATED — indexes are created in database_logic.init_db()"""
    # All database indexes are now created in database_logic.init_db()
    pass

ensure_indexes()
cleanup_old_tokens()


# ── Role helpers ───────────────────────────────────────────────────────────────

def current_user():
    return session.get("user_data")

def require_login():
    if not current_user():
        return redirect(url_for("login"))
    return None

def require_role(*roles):
    user = current_user()
    if not user:
        return redirect(url_for("login"))
    if user["role"] not in roles:
        return "❌ Access denied.", 403
    return None


# ─────────────────────────────────────────────
# LOGIN / LOGOUT
# ─────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = authenticate_user(username, password)
        if user:
            session["user_data"] = user
            return redirect(url_for("dashboard"))
        return render_template("login.html", error="❌ Invalid username or password.")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@app.route("/dashboard")
def dashboard():
    redir = require_login()
    if redir: return redir
    user         = current_user()
    subjects     = get_subjects_for_user(user)
    classes      = get_all_classes() if user["role"] in ["coordinator", "incharge"] else []
    all_users    = get_all_users()   if user["role"] == "coordinator" else []
    teachers     = get_teachers()
    incharges    = get_incharges()
    all_subjects = get_all_subjects() if user["role"] in ["coordinator", "incharge"] else []
    return render_template("dashboard.html",
        user=user,
        subjects=subjects,
        classes=classes,
        all_users=all_users,
        teachers=teachers,
        incharges=incharges,
        all_subjects=all_subjects
    )


# ─────────────────────────────────────────────
# USER MANAGEMENT  (coordinator only)
# ─────────────────────────────────────────────

@app.route("/add_user", methods=["POST"])
def add_user_route():
    err = require_role("coordinator")
    if err: return err
    username  = request.form.get("username", "").strip()
    password  = request.form.get("password", "").strip()
    role      = request.form.get("role", "")
    full_name = request.form.get("full_name", "").strip()
    class_id  = request.form.get("class_id") or None
    if class_id:
        class_id = int(class_id)
    add_user(username, password, role, full_name, class_id)
    return redirect(url_for("dashboard"))

@app.route("/delete_user/<int:uid>")
def delete_user_route(uid):
    err = require_role("coordinator")
    if err: return err
    delete_user(uid)
    return redirect(url_for("dashboard"))


# ─────────────────────────────────────────────
# CLASS MANAGEMENT  (coordinator only)
# ─────────────────────────────────────────────

@app.route("/add_class", methods=["POST"])
def add_class_route():
    err = require_role("coordinator")
    if err: return err
    name = request.form.get("class_name", "").strip()
    dept = request.form.get("department", "").strip()
    add_class(name, dept)
    return redirect(url_for("dashboard"))

@app.route("/delete_class/<int:cid>")
def delete_class_route(cid):
    err = require_role("coordinator")
    if err: return err
    delete_class(cid)
    return redirect(url_for("dashboard"))

@app.route("/assign_incharge", methods=["POST"])
def assign_incharge_route():
    err = require_role("coordinator")
    if err: return err
    class_id    = int(request.form.get("class_id"))
    incharge_id = int(request.form.get("incharge_id"))
    assign_incharge_to_class(class_id, incharge_id)
    return redirect(url_for("dashboard"))


# ─────────────────────────────────────────────
# SUBJECT MANAGEMENT  (coordinator + incharge)
# ─────────────────────────────────────────────

@app.route("/add_subject", methods=["POST"])
def add_subject_route():
    err = require_role("coordinator", "incharge")
    if err: return err
    user       = current_user()
    name       = request.form.get("subject_name", "").strip()
    class_id   = request.form.get("class_id") or user.get("class_id")
    teacher_id = request.form.get("teacher_id") or None
    if class_id:
        add_subject(name, int(class_id), int(teacher_id) if teacher_id else None)
    return redirect(url_for("dashboard"))

@app.route("/delete_subject/<int:sid>")
def delete_subject_route(sid):
    err = require_role("coordinator", "incharge")
    if err: return err
    delete_subject(sid)
    return redirect(url_for("dashboard"))

@app.route("/assign_teacher", methods=["POST"])
def assign_teacher_route():
    err = require_role("coordinator", "incharge")
    if err: return err
    user       = current_user()
    subject_id = int(request.form.get("subject_id"))
    teacher_id = int(request.form.get("teacher_id"))

    if user["role"] == "incharge":
        subject = get_subject_by_id(subject_id)
        if not subject or subject["class_id"] != user["class_id"]:
            return "❌ Access denied — subject not in your class.", 403

    assign_teacher_to_subject(subject_id, teacher_id)
    return redirect(url_for("dashboard"))


# ─────────────────────────────────────────────
# STUDENT MANAGEMENT
# ─────────────────────────────────────────────

@app.route("/students")
def show_students():
    redir = require_login()
    if redir: return redir
    user    = current_user()
    classes = get_all_classes()

    class_id = request.args.get("class_id")
    if class_id:
        class_id = int(class_id)
        students = get_students_by_class(class_id)
    elif user["role"] == "coordinator":
        students = get_all_students()
        class_id = None
    else:
        class_id = user.get("class_id")
        students = get_students_by_class(class_id) if class_id else []

    return render_template("students.html",
        user=user,
        students=students,
        classes=classes,
        selected_class_id=class_id
    )

@app.route("/add_student", methods=["POST"])
def add_student_route():
    err = require_role("coordinator", "incharge")
    if err: return err
    roll     = request.form.get("roll", "").strip()
    name     = request.form.get("name", "").strip()
    class_id = request.form.get("class_id", "")
    if not roll or not name or not class_id:
        return redirect(url_for("show_students"))
    add_student(roll, name, int(class_id))
    return redirect(url_for("show_students", class_id=class_id))

@app.route("/delete_student/<roll>")
def delete_student_route(roll):
    err = require_role("coordinator", "incharge")
    if err: return err
    class_id = request.args.get("class_id", "")
    delete_student(roll)
    return redirect(url_for("show_students", class_id=class_id))


# ─────────────────────────────────────────────
# QR GENERATION
# ─────────────────────────────────────────────

def _get_lan_ip():
    """Detect this machine's outbound LAN IP (e.g. 10.x.x.x or 192.168.x.x).
    Works by opening a UDP socket to an external address — no data is sent."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))          # connect to Google DNS (no data sent)
        lan_ip = s.getsockname()[0]
        s.close()
        return lan_ip
    except Exception:
        return "127.0.0.1"                  # fallback


def _get_base_url():
    """Return the server's base URL for embedding in QR codes.
    - localhost/127.0.0.1: swap in the LAN IP so students on the same network can reach the app.
    - Any other host (public domain, LAN IP already in the URL): use the request host as-is."""
    host = request.host.split(":")[0]
    if host in ("localhost", "127.0.0.1"):
        port = request.host.split(":")[-1] if ":" in request.host else "5000"
        return f"http://{_get_lan_ip()}:{port}"
    scheme = "https" if request.is_secure else "http"
    return f"{scheme}://{request.host}"


@app.route("/generate_qr/<int:subject_id>")
def generate_qr_route(subject_id):
    redir = require_login()
    if redir: return redir
    user    = current_user()
    subject = get_subject_by_id(subject_id)
    if not subject:
        return "❌ Subject not found."

    if user["role"] == "teacher":
        my_subjects = get_subjects_for_user(user)
        if not any(s["id"] == subject_id for s in my_subjects):
            return "❌ Access denied — not your subject.", 403

    base_url   = _get_base_url()
    session_id, qr_image = generate_qr(subject["id"], subject["name"], base_url=base_url)
    session["active_session_id"] = session_id
    session["active_subject_id"] = str(subject["id"])

    return render_template("qr_display.html",
        subject=subject,
        session_id=session_id,
        qr_image=qr_image
    )


@app.route("/refresh_qr/<int:subject_id>")
def refresh_qr(subject_id):
    redir = require_login()
    if redir: return jsonify({"error": "Not logged in"}), 401
    subject = get_subject_by_id(subject_id)
    if not subject:
        return jsonify({"error": "Subject not found"}), 404

    base_url   = _get_base_url()
    session_id, qr_image = generate_qr(subject["id"], subject["name"], base_url=base_url)
    session["active_session_id"] = session_id
    session["active_subject_id"] = str(subject["id"])

    return jsonify({"session_id": session_id, "qr_image": qr_image})


# ─────────────────────────────────────────────
# DEBUG: Attendance check endpoint
# ─────────────────────────────────────────────

@app.route("/debug/attendance_check")
def debug_attendance_check():
    """Debug endpoint to check all attendance records in database"""
    redir = require_login()
    if redir: return redir
    try:
        from database_logic import get_db
        conn = get_db()
        rows = conn.run("SELECT COUNT(*) FROM attendance")
        total_count = rows[0][0] if rows else 0
        
        rows = conn.run("SELECT subject_id, session_id, roll, name, timestamp FROM attendance ORDER BY timestamp DESC LIMIT 20")
        records = [{"subject_id": r[0], "session_id": r[1], "roll": r[2], "name": r[3], "timestamp": r[4]} for r in rows]
        conn.close()
        
        return jsonify({"total_records": total_count, "recent_records": records})
    except Exception as e:
        print(f"[ERROR] debug_attendance_check: {e}", flush=True)
        return jsonify({"error": str(e)}), 500

# ─────────────────────────────────────────────
# SCAN COUNT (polled by qr_display.html every 10 s)
# ─────────────────────────────────────────────

@app.route("/scan_count/<int:subject_id>/<session_id>")
def scan_count(subject_id, session_id):
    redir = require_login()
    if redir: return jsonify({"error": "Not logged in"}), 401
    try:
        # Import the database function from database_logic to use PostgreSQL
        from database_logic import get_db
        conn = get_db()
        print(f"[DEBUG] Querying attendance: subject_id={subject_id}, session_id={session_id}", flush=True)
        rows = conn.run(
            "SELECT COUNT(DISTINCT roll) FROM attendance "
            "WHERE subject_id = :sid AND session_id = :sess",
            sid=subject_id, sess=session_id
        )
        count = rows[0][0] if rows else 0
        print(f"[DEBUG] Query result: count={count}, rows={rows}", flush=True)
        conn.close()
        return jsonify({"count": count})
    except Exception as e:
        print(f"[ERROR] scan_count failed: {e}", flush=True)
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# SCAN (student side)
# ─────────────────────────────────────────────

@app.route("/scan")
def scan():
    subject_id = request.args.get("subject_id", "").strip()
    session_id = request.args.get("session_id", "").strip()

    print(f"[DEBUG] /scan endpoint hit: subject_id={subject_id}, session_id={session_id}", flush=True)

    subject = get_subject_by_id(int(subject_id)) if subject_id else None
    if not subject:
        print(f"[ERROR] /scan: Subject not found for subject_id={subject_id}", flush=True)
        return render_template("confirm.html",
            message="❌ Invalid QR code.",
            session_id="",
            success=False,
            student_name=None,
            roll=None,
            subject_name=None,
            class_name=None)

    device_token = issue_device_token(session_id)
    if not device_token:
        return render_template("confirm.html",
            message="❌ Invalid or expired session. Ask your teacher to refresh the QR.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=None,
            subject_name=subject["name"],
            class_name=subject.get("class_name", ""))

    session["device_token"]    = device_token
    session["scan_subject_id"] = subject_id
    session["scan_session_id"] = session_id

    print(f"[DEBUG] /scan: Successfully rendering scan.html with subject_id={subject_id}, session_id={session_id}, device_token={device_token[:16]}...", flush=True)

    return render_template("scan.html",
        subject=subject,
        session_id=session_id,
        device_token=device_token
    )


# ─────────────────────────────────────────────
# SUBMIT ATTENDANCE
# ─────────────────────────────────────────────

@app.route("/submit_attendance", methods=["POST"])
def submit_attendance():
    roll       = request.form.get("roll", "").strip().upper()
    subject_id = request.form.get("subject_id", "").strip()
    session_id = request.form.get("session_id", "").strip()
    ip_address = request.remote_addr

    # Prefer form-posted token; fall back to Flask session cookie
    device_token = request.form.get("device_token") or session.get("device_token")

    print(f"[DEBUG] submit_attendance called: roll={roll}, subject_id={subject_id}, session_id={session_id}, device_token={device_token[:16] if device_token else None}...", flush=True)

    if not device_token:
        return render_template("confirm.html",
            message="Invalid session. Please scan the QR code again.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=None,
            subject_name=None,
            class_name=None)

    # ── Step 1: Validate inputs ──────────────────────────────────────────────
    if not roll:
        return render_template("confirm.html",
            message="Please enter your roll number.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=None,
            subject_name=None,
            class_name=None)

    if not subject_id or not session_id:
        return render_template("confirm.html",
            message="Invalid QR code data. Please scan again.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=None,
            subject_name=None,
            class_name=None)

    subject = get_subject_by_id(int(subject_id))
    if not subject:
        return render_template("confirm.html",
            message="Subject not found. Contact your teacher.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=None,
            subject_name=None,
            class_name=None)

    # Grab subject details for use in all responses below
    subject_name = subject["name"]
    class_name   = subject.get("class_name", "")

    # ── Step 2: Validate roll number against the correct class ───────────────
    if not roll_exists(roll, subject["class_id"]):
        return render_template("confirm.html",
            message=f"Roll number '{roll}' not found in this class. Please check and scan again.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=roll,
            subject_name=subject_name,
            class_name=class_name)

    # ── Step 3: Check if this roll already marked attendance this session ────
    already, dup_reason = has_already_submitted(session_id, roll=roll)
    if already:
        return render_template("confirm.html",
            message="Attendance already marked for this roll number in this session.",
            session_id=session_id,
            success=False,
            student_name=None,
            roll=roll,
            subject_name=subject_name,
            class_name=class_name)

    # ── Step 4: Atomically consume the one-time device token ─────────────────
    allowed, reason = consume_device_token(session_id, device_token)
    if not allowed:
        error_messages = {
            "invalid_session":    "Session expired. Ask your teacher to refresh the QR code.",
            "unrecognized_token": "Invalid token. Please scan the QR code again.",
            "already_used":       "This device has already submitted for this session.",
        }
        return render_template("confirm.html",
            message=error_messages.get(reason, "Already submitted."),
            session_id=session_id,
            success=False,
            student_name=None,
            roll=roll,
            subject_name=subject_name,
            class_name=class_name)

    # ── Step 5: Insert attendance record ─────────────────────────────────────
    name   = get_student_name(roll) or roll
    marked = mark_attendance(
        subject["id"], subject["name"], subject["class_id"],
        session_id, roll, name, device_token, ip_address
    )

    session.pop("device_token", None)

    if not marked:
        # UNIQUE(session_id, roll) constraint fired — already exists
        return render_template("confirm.html",
            message="Attendance already recorded for this roll number.",
            session_id=session_id,
            success=False,
            student_name=name,
            roll=roll,
            subject_name=subject_name,
            class_name=class_name)

    # ── Step 6: Success ───────────────────────────────────────────────────────
    return render_template("confirm.html",
        message=f"Attendance marked for {name} ({roll})",
        session_id=session_id,
        success=True,
        student_name=name,
        roll=roll,
        subject_name=subject_name,
        class_name=class_name
    )


# ─────────────────────────────────────────────
# VIEW ATTENDANCE
# ─────────────────────────────────────────────

@app.route("/view_attendance/<int:subject_id>")
def view_attendance(subject_id):
    redir = require_login()
    if redir: return redir

    subject = get_subject_by_id(subject_id)
    if not subject:
        return "❌ Subject not found."

    students_data, date_cols = get_attendance_summary(subject_id, subject["class_id"])
    records = get_attendance_records(subject_id)

    return render_template("view_attendance.html",
        subject=subject,
        students_data=students_data,
        date_cols=date_cols,
        all_dates=date_cols,
        records=records,
        total_students=len(students_data),
        message="" if students_data else "⚠️ No attendance records found for this subject yet."
    )


# ─────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────

@app.route("/report/<int:subject_id>")
def report(subject_id):
    redir = require_login()
    if redir: return redir
    user    = current_user()
    subject = get_subject_by_id(subject_id)
    if not subject:
        return "❌ Subject not found."

    if user["role"] == "teacher":
        my_subjects = get_subjects_for_user(user)
        if not any(s["id"] == subject_id for s in my_subjects):
            return "❌ Access denied."
    elif user["role"] == "incharge":
        if subject["class_id"] != user["class_id"]:
            return "❌ Access denied."

    from datetime import datetime as dt
    import pandas as pd

    final_df, date_cols = generate_report_for_subject(subject_id, subject["class_id"])

    os.makedirs("exports", exist_ok=True)
    month_str = dt.now().strftime("%B_%Y")
    filename  = f"{_safe_name(subject['class_name'])}_{_safe_name(subject['name'])}_{month_str}.xlsx"
    filepath  = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Attendance", index=False)

    _apply_colors(filepath, date_cols)
    return redirect(url_for("download_export", filename=filename))


@app.route("/report/class/<int:class_id>")
def report_class(class_id):
    redir = require_login()
    if redir: return redir
    user = current_user()
    if user["role"] == "teacher":
        return "❌ Access denied."
    if user["role"] == "incharge" and user["class_id"] != class_id:
        return "❌ Access denied."

    from datetime import datetime as dt
    import pandas as pd
    from database_logic import get_db

    conn = get_db()
    subs = conn.run("SELECT id, name FROM subjects WHERE class_id = :c", c=class_id)
    cls_row = conn.run("SELECT name FROM classes WHERE id = :c", c=class_id)
    conn.close()

    if not subs:
        return "⚠️ No subjects found for this class."

    class_name = cls_row[0][0] if cls_row else f"Class_{class_id}"
    os.makedirs("exports", exist_ok=True)
    filename = f"{_safe_name(class_name)}_Full_Report_{dt.now().strftime('%B_%Y')}.xlsx"
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sub_id, sub_name in subs:
            final_df, date_cols = generate_report_for_subject(sub_id, class_id)
            sheet_name = sub_name[:31]
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    _apply_colors(filepath)
    return redirect(url_for("download_export", filename=filename))


@app.route("/report/all")
def report_all():
    err = require_role("coordinator")
    if err: return err

    from datetime import datetime as dt
    import pandas as pd
    from database_logic import get_db

    conn = get_db()
    classes = conn.run("SELECT id, name FROM classes ORDER BY name")
    conn.close()

    os.makedirs("exports", exist_ok=True)
    filename = f"All_Classes_Report_{dt.now().strftime('%B_%Y')}.xlsx"  # no user data in this name
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for cls_id, cls_name in classes:
            conn = get_db()
            subs = conn.run("SELECT id, name FROM subjects WHERE class_id = :c", c=cls_id)
            conn.close()
            for sub_id, sub_name in subs:
                sheet = f"{cls_name}-{sub_name}"[:31]
                final_df, _ = generate_report_for_subject(sub_id, cls_id)
                final_df.to_excel(writer, sheet_name=sheet, index=False)

    _apply_colors(filepath)
    return redirect(url_for("download_export", filename=filename))


def _safe_name(text):
    """Strip characters that are unsafe in filenames on Linux/Windows."""
    return "".join(c if c.isalnum() or c in " _-" else "_" for c in str(text)).strip()


def _apply_colors(filepath, date_cols=None):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(filepath)
    green       = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red         = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    yellow      = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        header_vals = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col_name = header_vals[cell.column - 1] if cell.column <= len(header_vals) else ""
                if cell.value == "P":
                    cell.fill      = green
                    cell.alignment = Alignment(horizontal="center")
                elif cell.value == "A":
                    cell.fill      = red
                    cell.alignment = Alignment(horizontal="center")
                elif col_name == "Attendance %":
                    try:
                        pct = float(cell.value)
                        if pct >= 75:
                            cell.fill = green
                        elif pct >= 60:
                            cell.fill = yellow
                        else:
                            cell.fill = red
                    except (TypeError, ValueError):
                        pass
                    cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len    = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 20)

    wb.save(filepath)


# ─────────────────────────────────────────────
# MISC
# ─────────────────────────────────────────────

@app.route("/download/<filename>")
def download_export(filename):
    from flask import send_from_directory
    exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
    return send_from_directory(exports_dir, filename, as_attachment=True)

@app.route("/refresh_students")
def refresh_students():
    err = require_role("coordinator", "incharge")
    if err: return err
    user     = current_user()
    class_id = request.args.get("class_id") or user.get("class_id")
    if class_id:
        seed_students_from_excel(class_id=int(class_id))
    else:
        seed_all_classes()
    return redirect(url_for("dashboard"))

@app.route("/get_name", methods=["POST"])
def get_name():
    roll = request.form["roll"].strip().upper()
    name = get_student_name(roll)
    return name if name else "❌ Not found"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)