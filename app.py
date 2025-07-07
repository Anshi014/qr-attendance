from database_logic import init_db, seed_students_from_excel
from database_logic import load_student_list
from flask import Flask, render_template, request, redirect, url_for, session
from database_logic import roll_exists
from users import users
    # Redirect to dashboard
from qr_generator import generate_qr
import os
app = Flask(__name__)
app.secret_key = "secretkey123"
init_db() #this creates the attendance table if not already there
seed_students_from_excel()

# Needed to manage sessions
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username in users and users[username]["password"] == password:
            session["user"] = username
            return redirect(url_for("dashboard"))

        return "❌ Invalid login. Try again."

    return render_template("login.html")
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))

    user = users[session["user"]]
    return render_template(
        "dashboard.html",
        username=session["user"],
        role=user["role"],
        subjects=user["subjects"]
    )
@app.route("/generate_qr/<subject>")
def generate_qr_route(subject):
    if "user" not in session:
        return redirect(url_for("login"))

    session_id, subject, qr_image = generate_qr(subject)
    
    user = users.get(session["user"])
    import sqlite3
    from datetime import date

    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT subject, COUNT(DISTINCT roll) FROM attendance
        WHERE DATE(timestamp) = ?
        GROUP BY subject
    """, (date.today().isoformat(),))
    daily_counts = dict(cursor.fetchall())
    conn.close()

    return redirect(url_for("qr_display", subject=subject, session_id=session_id, qr_image=qr_image))

@app.route("/scan")
def scan():
    subject = request.args.get("subject", "")
    session_id = request.args.get("session_id", "")
    roll = request.args.get("roll", "").strip().upper()

    key = f"submitted_{subject}_{session_id}_{roll}"
    if session.get(key):
        return render_template("confirm.html", message="❌ This device already submitted for this student.")

    return render_template("scan.html", subject=subject, session_id=session_id, roll=roll)
from database_logic import has_already_submitted, mark_attendance
@app.route("/submit_attendance", methods=["POST"])
def submit_attendance():
    roll = request.form["roll"].strip().upper()
    subject = request.form["subject"]
    session_id = request.form["session_id"]
    device_id = request.form.get("device_id")
    ip_address = request.remote_addr
    
    print("📲 Roll:", roll)
    print("📶 Device ID:", device_id)
    print("🌐 IP Address:", ip_address)


    # Check if roll exists
    if not roll_exists(roll):
        return render_template("confirm.html", message="❌ Roll number not found.")

    # Get name from DB
    import sqlite3
    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE roll = ?", (roll,))
    result = cursor.fetchone()
    conn.close()
    name = result[0] if result else "Unknown"

    # ✅ BLOCK: Same roll cannot mark twice
    if has_already_submitted(subject, session_id, device_id=device_id, roll=roll):
        return render_template("confirm.html", message="❌ This device already marked attendance for this subject.")

    # ✅ BLOCK: Same device/IP can't mark for multiple rolls
    if has_already_submitted(subject=subject, ip_address=ip_address):
        return render_template("confirm.html", message="❌ This IP already submitted attendance for this subject.")

    # ✅ Save attendance
    mark_attendance(subject, session_id, roll, name, device_id, ip_address)
    return render_template("confirm.html", message=f"✅ Attendance marked for {name} ({roll})")

@app.route("/export/<subject>")
def export_subject(subject):
    import pandas as pd
    import sqlite3

    if "user" not in session:
        return redirect(url_for("login"))
    
    user = users.get(session["user"])
    if user["role"] not in ["incharge","cr"]:
        return "Access denied."

    conn = sqlite3.connect("attendance.db")
    df = pd.read_sql_query("SELECT * FROM attendance ", conn)
    conn.close()

    export_folder = "exports"
    os.makedirs(export_folder, exist_ok=True)
    filename = f"{subject}_attendance.xlsx"
    filepath = os.path.join(export_folder, filename)
    df.to_excel(filepath, index=False)

    return redirect(url_for("download_export", filename=filename))

@app.route("/download/<filename>")
def download_export(filename):
    from flask import send_from_directory
    return send_from_directory("exports", filename, as_attachment=True)

@app.route("/students")
def show_students():
    df = load_student_list()
    if df is None:
        return "❌ Could not load student list."

    return df.to_html(index=False)

@app.route("/report/<subject>/<session_id>")
def generate_report(subject, session_id):
    import pandas as pd
    import sqlite3
    from datetime import datetime
    import os

    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()

    # Print student list
    cursor.execute("SELECT roll, name FROM students")
    students = cursor.fetchall()
    print("📋 Students:", students)

    cursor.execute("""
        SELECT roll, timestamp FROM attendance
        WHERE subject = ? AND session_id = ?
    """, (subject, session_id))
    present_data = {row[0].strip().upper(): row[1] for row in cursor.fetchall()}
    present_rolls = set(present_data.keys())
    print("🟢 Present data:", present_data)

    conn.close()

    if present_data:
        first_timestamp = list(present_data.values())[0]
        try:
            date_str = datetime.strptime(first_timestamp, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        except:
            date_str = datetime.now().strftime("%Y-%m-%d")
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    report = []
    for roll, name in students:
        roll_clean = roll.strip().upper()
        status = "P" if roll_clean in present_rolls else "A"
        report.append({
            "Roll": roll,
            "Name": name,
            date_str: status
        })

    df = pd.DataFrame(report)
    export_folder = "exports"
    os.makedirs(export_folder, exist_ok=True)
    filename = f"{subject}_{session_id}_report.xlsx"
    filepath = os.path.join(export_folder, filename)
    df.to_excel(filepath, index=False)

    return redirect(url_for("download_export", filename=filename))

@app.route("/get_name", methods=["POST"])
def get_name():
    roll = request.form["roll"].strip().upper()
    import sqlite3
    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM students WHERE roll = ?", (roll,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else "❌ Not found"

@app.route("/refresh_students")
def refresh_students():
    from database_logic import seed_students_from_excel
    seed_students_from_excel()
    return "✅ Student list refreshed from Excel!"

@app.route("/full_report/<subject>")
def full_report(subject):
    if "user" not in session:
        return redirect(url_for("login"))

    user = users.get(session["user"])
    if user["role"] not in ["incharge", "cr"]:
        return "❌ Access denied. Only CR/Incharge can view full report."

    import pandas as pd
    import sqlite3
    import os

    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()

    # Get full student list
    cursor.execute("SELECT roll, name FROM students")
    students = cursor.fetchall()

    # Get all attendance data for this subject
    cursor.execute("""
        SELECT roll, DATE(timestamp) as date
        FROM attendance
        WHERE subject = ?
    """, (subject,))
    records = cursor.fetchall()
    conn.close()

    # Convert attendance to DataFrame
    df = pd.DataFrame(records, columns=["Roll", "Date"])
    df["Status"] = "P"

    # Pivot date-wise report
    pivot = df.pivot_table(index="Roll", columns="Date", values="Status", aggfunc="first", fill_value="A")

    # Merge with student names
    student_df = pd.DataFrame(students, columns=["Roll", "Name"])
    final_df = pd.merge(student_df, pivot, on="Roll", how="left").fillna("A")

    # ✅ Add total 'P' and % attendance
    status_cols = final_df.columns[2:]  # skip Roll and Name
    final_df["Total Present"] = final_df[status_cols].apply(lambda row: sum(x == "P" for x in row), axis=1)
    final_df["% Attendance"] = (final_df["Total Present"] / len(status_cols) * 100).round(2)

    # ⬇️ Insert debug prints here
    print("▶️ Attendance Records:", records)
    print("▶️ Pivot Columns:", pivot.columns.tolist())
    print("▶️ Final Report Preview:\n", final_df.head())

    # Export
    os.makedirs("exports", exist_ok=True)
    from datetime import datetime
    month_str = datetime.now().strftime("%B")
    filename = f"{subject}_{month_str}_report.xlsx"

    final_df.to_excel(os.path.join("exports", filename), index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(os.path.join("exports", filename))
    ws = wb.active  
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=ws.max_column - 2):
        for cell in row:
            if cell.value == "P":
                cell.fill = green
            elif cell.value == "A":
                cell.fill = red

    wb.save(os.path.join("exports", filename))

    return redirect(url_for("download_export", filename=filename))

@app.route("/debug_attendance")
def debug_attendance():
    import sqlite3
    conn = sqlite3.connect("attendance.db")
    cursor = conn.cursor()
    cursor.execute("SELECT roll, subject, session_id, timestamp FROM attendance")
    rows = cursor.fetchall()
    conn.close()
    return "<br>".join([str(row) for row in rows])

@app.route("/test")
def test_page():
    return "✅ Server is running!"

@app.route("/qr_display/<subject>/<session_id>/<qr_image>")
def qr_display(subject, session_id, qr_image):
    return render_template("qr_display.html",
        subject=subject,
        session_id=session_id,
        qr_image=qr_image
    )


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
